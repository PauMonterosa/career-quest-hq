import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session
from ..models import ActionTask, ApplicationDocument, EmailDraft, MasterProgramme, TFGOpportunity

log = logging.getLogger(__name__)


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def safe(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value if value not in ("", None) else None


ALIASES = {
    "name": ["nom", "master", "programa", "program", "titol", "title", "document", "tasca", "possible_tfg"],
    "university": ["universitat", "universitat_consorci", "university", "centre"],
    "location": ["ubicacio", "ciutat", "location", "pais_mobilitat"],
    "status": ["estat", "status"],
    "score": ["puntuacio", "puntuacio_ponderada", "score", "nota"],
    "deadline": ["deadline", "termini", "data_limit", "finestra_orientativa"],
    "prerequisites": ["prerequisits", "prerequisites", "requisits", "punt_clau_risc"],
    "url": ["url", "web", "link", "font_oficial"],
    "centre": ["centre", "centre_grup", "centre_recerca", "institucio"],
    "supervisor": ["supervisor", "tutor", "investigador"],
    "topic": ["tema", "topic", "linia", "linia_recerca", "descripcio"],
    "contact": ["contacte", "contact", "email"],
    "category": ["categoria", "category", "area", "fase"],
    "priority": ["prioritat", "priority"],
    "due_date": ["data", "data_limit", "due_date", "termini"],
    "notes": ["notes", "observacions", "comentaris"],
    "subject": ["assumpte", "subject", "tema"],
    "recipient": ["destinatari", "recipient", "contacte", "email"],
    "body": ["cos", "cos_del_correu", "body", "missatge", "text"],
    "follow_up_date": ["seguiment", "follow_up", "data_seguiment"],
}


def pick(data: dict[str, Any], key: str) -> Any:
    for alias in ALIASES.get(key, [key]):
        if normalize(alias) in data:
            return safe(data[normalize(alias)])
    return None


@dataclass
class ImportSummary:
    workbook: str
    imported: dict[str, int] = field(default_factory=dict)
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


SHEET_MODELS = {
    "mapa_masters": MasterProgramme,
    "tfg_barcelona": TFGOpportunity,
    "pla_d_accio": ActionTask,
    "correus": EmailDraft,
    "documents": ApplicationDocument,
}


def _headers_and_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    for row_number, row in enumerate(rows, 1):
        values = [safe(v) for v in row]
        if sum(v is not None for v in values) >= 2:
            headers = [normalize(v) or f"column_{i + 1}" for i, v in enumerate(values)]
            for source_row, data_row in enumerate(rows, row_number + 1):
                yield source_row, {headers[i]: safe(v) for i, v in enumerate(data_row) if i < len(headers)}
            return


def _build(model, data, sheet, row):
    common = {"source_sheet": sheet, "source_row": row, "source_data": data}
    if model is MasterProgramme:
        return model(name=pick(data, "name") or f"Master row {row}", university=pick(data, "university"),
            location=pick(data, "location"), status=pick(data, "status"), score=str(pick(data, "score") or "") or None,
            deadline=pick(data, "deadline"), prerequisites=pick(data, "prerequisites"), url=pick(data, "url"), **common)
    if model is TFGOpportunity:
        return model(title=pick(data, "name") or pick(data, "topic") or f"TFG row {row}", centre=pick(data, "centre"),
            supervisor=pick(data, "supervisor"), topic=pick(data, "topic"), status=pick(data, "status"),
            contact=pick(data, "contact"), **common)
    if model is ActionTask:
        return model(title=pick(data, "name") or f"Task row {row}", category=pick(data, "category"),
            status=pick(data, "status"), priority=pick(data, "priority"), due_date=pick(data, "due_date"),
            notes=pick(data, "notes"), **common)
    if model is EmailDraft:
        return model(subject=pick(data, "subject") or f"Email row {row}", recipient=pick(data, "recipient"),
            body=pick(data, "body"), status=pick(data, "status"), follow_up_date=pick(data, "follow_up_date"),
            approval_required=True, **common)
    return model(name=pick(data, "name") or f"Document row {row}", status=pick(data, "status"),
        notes=pick(data, "notes"), **common)


def import_workbook(path: Path, db: Session, replace: bool = False) -> ImportSummary:
    summary = ImportSummary(workbook=str(path))
    if not path.exists():
        summary.errors.append(f"Workbook not found: {path}")
        return summary
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            model = SHEET_MODELS.get(normalize(sheet.title))
            if model is None:
                continue
            if replace:
                db.execute(delete(model))
            count = 0
            for source_row, data in _headers_and_rows(sheet) or []:
                if not any(v is not None for v in data.values()):
                    summary.skipped += 1
                    continue
                try:
                    db.add(_build(model, data, sheet.title, source_row))
                    count += 1
                except Exception as exc:
                    message = f"{sheet.title}!{source_row}: {exc}"
                    log.exception(message)
                    summary.errors.append(message)
            summary.imported[model.__tablename__] = count
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()
    return summary
