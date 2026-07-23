from openpyxl import Workbook, load_workbook
from app.models import ActionTask, MasterProgramme, TFGOpportunity
from app.services.excel_importer import import_workbook


def test_imports_expected_sheets_without_modifying_source(tmp_path, db):
    path = tmp_path / "plan.xlsx"
    book = Workbook()
    masters = book.active
    masters.title = "Mapa màsters"
    masters.append(["Programa", "Universitat / consorci", "Puntuació ponderada", "Estat", "Font oficial"])
    masters.append(["MSc Solar Energy", "UPC", 9, "Shortlist"])
    tfg = book.create_sheet("TFG Barcelona")
    tfg.append(["Centre / grup", "Línia", "Possible TFG", "Prioritat"])
    tfg.append(["IREC", "PV devices", "PV forecasting", "Molt alta"])
    tasks = book.create_sheet("Pla d'acció")
    tasks.append(["Fase", "Tasca", "Prioritat", "Data límit"])
    tasks.append(["Portafolis", "Prepare CV", "Alta", "2026-09-01"])
    book.save(path)
    before = path.read_bytes()

    summary = import_workbook(path, db)

    assert path.read_bytes() == before
    assert summary.errors == []
    assert db.query(MasterProgramme).one().source_row == 2
    assert db.query(MasterProgramme).one().university == "UPC"
    assert db.query(MasterProgramme).one().score == "9"
    assert db.query(TFGOpportunity).one().centre == "IREC"
    assert db.query(TFGOpportunity).one().title == "PV forecasting"
    assert db.query(ActionTask).one().priority == "Alta"
    assert db.query(ActionTask).one().title == "Prepare CV"
    assert load_workbook(path, read_only=True).sheetnames == ["Mapa màsters", "TFG Barcelona", "Pla d'acció"]


def test_missing_workbook_returns_summary(tmp_path, db):
    summary = import_workbook(tmp_path / "missing.xlsx", db)
    assert "Workbook not found" in summary.errors[0]
